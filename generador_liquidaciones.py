"""
Sistema de Generación de Liquidaciones por RUC y CAMPAÑA
Extrae datos de los archivos DetalleEmpresas por campaña y genera liquidaciones formateadas
Mapeo de campañas:
- DetalleEmpresas_Camp_717.xlsx -> PRESUNTA
- DetalleEmpresas_Camp_714.xlsx -> DEUDA REAL TOTAL
- DetalleEmpresas_Camp_713.xlsx -> REDIRECCIONAMIENTO
- DetalleEmpresas_Camp_709.xlsx -> PREJUDICIAL FLUJO

Incluye cálculo automático de mora e intereses usando factor_interes.xlsx
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
from pathlib import Path
from calculador_intereses import CalculadorIntereses

class GeneradorLiquidaciones:
    # Mapeo de archivos a campañas
    CAMPANAS = {
        "DetalleEmpresas_Camp_717.xlsx": "PRESUNTA",
        "DetalleEmpresas_Camp_714.xlsx": "DEUDA REAL TOTAL",
        "DetalleEmpresas_Camp_713.xlsx": "REDIRECCIONAMIENTO",
        "DetalleEmpresas_Camp_709.xlsx": "PREJUDICIAL FLUJO"
    }
    
    def __init__(self, base_path):
        self.base_path = base_path
        self.detalle_files = list(self.CAMPANAS.keys())
        self.formato_path = os.path.join(base_path, "formato.xlsx")
        
        # Cargar calculador de intereses
        self.calculador_intereses = CalculadorIntereses(base_path)
        self.calculador_intereses.establecer_dia()  # Usar día actual
        
        # Cargar datos de detalle por campaña
        self.datos_por_campana = {}  # Dict: campaña -> DataFrame
        self.datos_completos = None  # Todos los datos combinados
        self.rucs_disponibles = set()
        self.rucs_por_campana = {}  # Dict: (ruc, campaña) -> booleano
        self._cargar_datos()
    
    def _cargar_datos(self):
        """Carga todos los datos de los archivos de detalle organizados por campaña"""
        dfs = []
        
        for archivo, campana in self.CAMPANAS.items():
            ruta = os.path.join(self.base_path, archivo)
            print(f"Cargando {archivo} ({campana})...")
            
            df = pd.read_excel(ruta)
            # Agregar columna de campaña
            df['CAMPANA'] = campana
            
            self.datos_por_campana[campana] = df
            dfs.append(df)
            
            # Registrar RUCs por campaña
            rucs_campana = set(df['DOCUMENTO'].unique())
            for ruc in rucs_campana:
                self.rucs_por_campana[(ruc, campana)] = True
                self.rucs_disponibles.add(ruc)
        
        # Combinar todos los datos
        self.datos_completos = pd.concat(dfs, ignore_index=True)
        print(f"\nTotal de registros cargados: {len(self.datos_completos)}")
        print(f"RUCs únicos totales: {len(self.rucs_disponibles)}")
    
    def obtener_rucs(self):
        """Retorna lista de RUCs disponibles"""
        return sorted(list(self.rucs_disponibles))
    
    def obtener_campanas_ruc(self, ruc):
        """Retorna las campañas en que aparece un RUC específico"""
        campanas = []
        for campana in self.CAMPANAS.values():
            if (ruc, campana) in self.rucs_por_campana:
                campanas.append(campana)
        return campanas
    
    def filtrar_por_ruc_campana(self, ruc, campana):
        """Filtra los datos por RUC y CAMPAÑA específicos, con cálculo de mora"""
        if campana not in self.CAMPANAS.values():
            raise ValueError(f"Campaña {campana} no válida")
        
        if (ruc, campana) not in self.rucs_por_campana:
            raise ValueError(f"RUC {ruc} no encontrado en campaña {campana}")
        
        datos_ruc = self.datos_completos[
            (self.datos_completos['DOCUMENTO'] == ruc) & 
            (self.datos_completos['CAMPANA'] == campana)
        ].copy()
        
        # Agrupar por CUSSP (afiliado) y PROCESO (período)
        datos_ruc = datos_ruc.groupby(['CUSSP', 'OPERACION', 'RAZON_SOCIAL', 'AFILIADO']).agg({
            'FONDO_NOMINAL': 'sum',
            'SEGURO_NOMINAL': 'sum',
            'COMISION_NOMINAL': 'sum',
            'AFP_NOMINAL': 'sum',
            'TOTA_FONDO': 'sum'
        }).reset_index()
        
        # Calcular mora e intereses para cada fila
        if self.calculador_intereses.df_factor is not None:
            datos_ruc['MORA'] = 0.0
            datos_ruc['DEUDA_CON_MORA'] = datos_ruc['TOTA_FONDO']
            
            for idx, row in datos_ruc.iterrows():
                deuda_mora, mora, _ = self.calculador_intereses.calcular_deuda_con_mora(
                    row['TOTA_FONDO'], 
                    row['OPERACION']
                )
                datos_ruc.at[idx, 'MORA'] = mora
                datos_ruc.at[idx, 'DEUDA_CON_MORA'] = deuda_mora
        else:
            # Si no hay factores de interés, usar deuda como está
            datos_ruc['MORA'] = 0.0
            datos_ruc['DEUDA_CON_MORA'] = datos_ruc['TOTA_FONDO']
        
        return datos_ruc
    
    def filtrar_por_ruc(self, ruc):
        """Filtra los datos por RUC de todas las campañas (compatibilidad)"""
        if ruc not in self.rucs_disponibles:
            raise ValueError(f"RUC {ruc} no encontrado en los datos")
        
        datos_ruc = self.datos_completos[self.datos_completos['DOCUMENTO'] == ruc].copy()
        
        # Agrupar por CUSSP (afiliado) y PROCESO (período)
        datos_ruc = datos_ruc.groupby(['CUSSP', 'OPERACION', 'RAZON_SOCIAL', 'AFILIADO']).agg({
            'FONDO_NOMINAL': 'sum',
            'SEGURO_NOMINAL': 'sum',
            'COMISION_NOMINAL': 'sum',
            'AFP_NOMINAL': 'sum',
            'TOTA_FONDO': 'sum'
        }).reset_index()
        
        # Calcular mora e intereses
        if self.calculador_intereses.df_factor is not None:
            datos_ruc['MORA'] = 0.0
            datos_ruc['DEUDA_CON_MORA'] = datos_ruc['TOTA_FONDO']
            
            for idx, row in datos_ruc.iterrows():
                deuda_mora, mora, _ = self.calculador_intereses.calcular_deuda_con_mora(
                    row['TOTA_FONDO'],
                    row['OPERACION']
                )
                datos_ruc.at[idx, 'MORA'] = mora
                datos_ruc.at[idx, 'DEUDA_CON_MORA'] = deuda_mora
        else:
            datos_ruc['MORA'] = 0.0
            datos_ruc['DEUDA_CON_MORA'] = datos_ruc['TOTA_FONDO']
        
        return datos_ruc
    
    def generar_liquidacion(self, ruc, campana=None, razon_social=None, direccion="", fecha_pago=None):
        """Genera un archivo de liquidación para un RUC específico y campaña
        
        Args:
            ruc: RUC del deudor
            campana: Nombre de la campaña. Si es None, genera para todas las campañas del RUC
            razon_social: Nombre de la empresa (optional)
            direccion: Dirección de la empresa (optional)
            fecha_pago: Fecha de pago en formato YYYY-MM-DD (optional)
        """
        
        # Si no se especifica campaña, generar para todas
        if campana is None:
            campanas = self.obtener_campanas_ruc(ruc)
            if not campanas:
                print(f"No hay datos para el RUC {ruc}")
                return None
            
            archivos_generados = []
            for camp in campanas:
                archivo = self.generar_liquidacion(ruc, campana=camp, razon_social=razon_social, 
                                                  direccion=direccion, fecha_pago=fecha_pago)
                if archivo:
                    archivos_generados.append(archivo)
            return archivos_generados
        
        # Obtener datos del RUC en la campaña específica
        datos_ruc = self.filtrar_por_ruc_campana(ruc, campana)
        
        if datos_ruc.empty:
            print(f"No hay datos para el RUC {ruc} en campaña {campana}")
            return None
        
        # Obtener primera fila para información de la empresa
        datos_origenes = self.datos_completos[
            (self.datos_completos['DOCUMENTO'] == ruc) & 
            (self.datos_completos['CAMPANA'] == campana)
        ]
        primera_fila = datos_origenes.iloc[0]
        
        if razon_social is None:
            razon_social = primera_fila.get('RAZON_SOCIAL', 'N/A')
        
        if fecha_pago is None:
            fecha_pago = datetime.now().strftime('%Y-%m-%d')
        
        # Crear archivo basado en el formato
        wb = load_workbook(self.formato_path)
        ws = wb.active
        
        # Cambiar nombre de la hoja
        ws.title = f"LIQ_{ruc}_{campana[:3].upper()}"
        
        # Actualizar datos de encabezado
        ws['B7'] = "GI CORONADO"  # Empresa emisora
        ws['B13'] = f"Razón Social : {razon_social}"
        ws['G13'] = f"Ruc : {ruc}"
        ws['B14'] = f"Direccion : {direccion}"
        ws['B15'] = f"Fecha de pago : {fecha_pago}"
        
        # Agregar información de campaña (nueva fila)
        ws['B16'] = f"Campaña : {campana}"
        ws['B16'].font = Font(bold=True, size=10)
        
        # Limpiar filas de detalle (18-71)
        for row in range(19, 72):
            for col in range(2, 13):  # Columnas B a L
                ws.cell(row=row, column=col).value = None
        
        # Insertar datos de liquidación
        total_fondo = 0
        total_administradora = 0
        total_general = 0
        
        # Escribir detalles
        start_row = 19
        for idx, row in datos_ruc.iterrows():
            excel_row = start_row + idx
            
            # Usar DEUDA_CON_MORA en cálculos
            deuda_con_mora = row['DEUDA_CON_MORA']
            deuda_base = row['TOTA_FONDO']
            mora = row['MORA']
            
            # RUC
            ws.cell(row=excel_row, column=2).value = row['CUSSP']
            # PERIODOS
            ws.cell(row=excel_row, column=3).value = row['OPERACION']
            # FONDO
            ws.cell(row=excel_row, column=4).value = row['FONDO_NOMINAL']
            # ADMINISTRADORA
            ws.cell(row=excel_row, column=5).value = row['COMISION_NOMINAL']
            # FACTOR DE INTERES (basado en mora)
            factor = (deuda_con_mora / deuda_base) if deuda_base > 0 else 1.0
            ws.cell(row=excel_row, column=6).value = round(factor, 4)
            # INTERES FONDO (mora calculada)
            ws.cell(row=excel_row, column=7).value = round(mora, 2)
            # INTERES ADMINISTRADORA
            interes_admin = row['SEGURO_NOMINAL'] + row['AFP_NOMINAL']
            ws.cell(row=excel_row, column=8).value = round(interes_admin, 2)
            # TOTAL FONDO (con mora)
            ws.cell(row=excel_row, column=9).value = round(deuda_con_mora, 2)
            # TOTAL ADMINISTRADORA
            total_admin = row['COMISION_NOMINAL'] + interes_admin
            ws.cell(row=excel_row, column=10).value = round(total_admin, 2)
            # TOTAL
            total_row = deuda_con_mora + total_admin
            ws.cell(row=excel_row, column=11).value = round(total_row, 2)
            # AFILIADO (nombre)
            ws.cell(row=excel_row, column=12).value = row['RAZON_SOCIAL']
            
            total_fondo += deuda_con_mora
            total_administradora += total_admin
            total_general += total_row
        
        # Actualizar totales
        ws['F74'] = round(total_fondo, 2)
        ws['F75'] = round(total_fondo * 0.15, 2)  # 15% gastos cobranza
        ws['F76'] = round((total_fondo * 0.15) * 0.18, 2)  # IGV 18%
        gastos_admin = round(total_fondo * 0.15, 2) + round((total_fondo * 0.15) * 0.18, 2)
        ws['F77'] = gastos_admin
        ws['F78'] = round(total_fondo + gastos_admin, 2)
        ws['K78'] = gastos_admin
        
        # Guardar archivo con nombre que incluya la campaña
        campana_abrev = campana.replace(" ", "_").upper()[:10]
        nombre_archivo = f"LIQUIDACION_{ruc}_{campana_abrev}_{fecha_pago.replace('-', '')}.xlsx"
        ruta_salida = os.path.join(self.base_path, "LIQUIDACIONES_GENERADAS")
        
        if not os.path.exists(ruta_salida):
            os.makedirs(ruta_salida)
        
        ruta_completa = os.path.join(ruta_salida, nombre_archivo)
        wb.save(ruta_completa)
        
        print(f"\n[OK] Liquidacion generada: {nombre_archivo}")
        print(f"  - RUC: {ruc}")
        print(f"  - Campana: {campana}")
        print(f"  - Razon Social: {razon_social}")
        print(f"  - Total Deuda: S/. {round(total_fondo, 2)}")
        print(f"  - Gastos Administrativos: S/. {gastos_admin}")
        print(f"  - Total con Gastos: S/. {round(total_fondo + gastos_admin, 2)}")
        
        return ruta_completa


def main():
    base_path = r"C:\Users\USUARIO\Desktop\REPORTE MENSUAL WORLDTEL\GENERACION DE LIQUIDACIONES"
    
    # Crear generador
    gen = GeneradorLiquidaciones(base_path)
    
    # Obtener RUCs disponibles
    rucs = gen.obtener_rucs()
    print(f"\n{'='*80}")
    print(f"RUCS DISPONIBLES PARA GENERAR LIQUIDACIONES")
    print(f"{'='*80}")
    print(f"Total de RUCs únicos: {len(rucs)}\n")
    
    # Contar casos únicos por campaña
    total_casos = 0
    for campana in GeneradorLiquidaciones.CAMPANAS.values():
        casos = len([ruc for ruc in rucs if (ruc, campana) in gen.rucs_por_campana])
        print(f"  {campana}: {casos} casos")
        total_casos += casos
    
    print(f"\nTotal de casos (RUC x Campaña): {total_casos}")
    
    # Mostrar primeros 20 RUCs
    print(f"\nPrimeros RUCs disponibles:")
    for i, ruc in enumerate(rucs[:20]):
        campanas = gen.obtener_campanas_ruc(ruc)
        print(f"{i+1:3d}. RUC: {ruc} - Campañas: {', '.join(campanas)}")
    
    if len(rucs) > 20:
        print(f"... y {len(rucs) - 20} RUCs más")
    
    # Ejemplo: Generar liquidación para el primer RUC
    if rucs:
        print(f"\n{'='*80}")
        print("GENERANDO LIQUIDACIONES DE EJEMPLO")
        print(f"{'='*80}")
        
        ruc_ejemplo = rucs[0]
        print(f"\nRUC: {ruc_ejemplo}")
        
        campanas_ejemplo = gen.obtener_campanas_ruc(ruc_ejemplo)
        print(f"Campaña(s) disponible(s): {', '.join(campanas_ejemplo)}")
        
        # Generar para la primera campaña
        if campanas_ejemplo:
            gen.generar_liquidacion(ruc_ejemplo, campana=campanas_ejemplo[0])


if __name__ == "__main__":
    main()
